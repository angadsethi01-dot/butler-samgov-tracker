from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORTS_DIR


def write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def write_excel(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Butler Opportunities", index=False)
        ws = writer.book["Butler Opportunities"]
        ws.freeze_panes = "A2"
        fills = {
            "A": PatternFill("solid", fgColor="D9EAD3"),
            "RFP": PatternFill("solid", fgColor="D9EAF7"),
            "Sources": PatternFill("solid", fgColor="FFF2CC"),
            "DueSoon": PatternFill("solid", fgColor="F4CCCC"),
            "Vehicle": PatternFill("solid", fgColor="EADCF8"),
            "Partial": PatternFill("solid", fgColor="FCE5CD"),
        }
        headers = {cell.value: cell.column for cell in ws[1]}
        for row_idx in range(2, ws.max_row + 1):
            fit = ws.cell(row=row_idx, column=headers.get("Fit Category", 1)).value
            stage = ws.cell(row=row_idx, column=headers.get("Opportunity Stage", 1)).value or ""
            days = ws.cell(row=row_idx, column=headers.get("Days Until Due", 1)).value
            vehicle = ws.cell(row=row_idx, column=headers.get("SeaPort / Vehicle Required", 1)).value or ""
            refresh = ws.cell(row=row_idx, column=headers.get("Refresh Status", 1)).value or ""
            row_fill = None
            if "Partial" in str(refresh):
                row_fill = fills["Partial"]
            elif "SeaPort" in str(vehicle) or "IDIQ" in str(vehicle):
                row_fill = fills["Vehicle"]
            elif isinstance(days, int) and days <= 7:
                row_fill = fills["DueSoon"]
            elif "Sources" in str(stage):
                row_fill = fills["Sources"]
            elif "RFP" in str(stage) or "Solicitation" in str(stage):
                row_fill = fills["RFP"]
            elif fit == "A":
                row_fill = fills["A"]
            if row_fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 55)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")


def create_exports(
    main_df: pd.DataFrame,
    rejected_df: pd.DataFrame,
    failed_df: pd.DataFrame,
    refresh_log_df: pd.DataFrame,
    partial_df: pd.DataFrame | None = None,
    exports_dir: Path = EXPORTS_DIR,
) -> dict[str, Path]:
    exports_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "main_results.csv": exports_dir / "main_results.csv",
        "main_results.xlsx": exports_dir / "main_results.xlsx",
        "rejected_results.csv": exports_dir / "rejected_results.csv",
        "partial_refresh_results.csv": exports_dir / "partial_refresh_results.csv",
        "failed_queries.csv": exports_dir / "failed_queries.csv",
        "refresh_log.csv": exports_dir / "refresh_log.csv",
    }
    write_csv(main_df, paths["main_results.csv"])
    write_excel(main_df, paths["main_results.xlsx"])
    write_csv(rejected_df, paths["rejected_results.csv"])
    write_csv(partial_df if partial_df is not None else main_df, paths["partial_refresh_results.csv"])
    write_csv(failed_df, paths["failed_queries.csv"])
    write_csv(refresh_log_df, paths["refresh_log.csv"])
    return paths

